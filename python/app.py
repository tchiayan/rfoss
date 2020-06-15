from openpyxl import load_workbook
import pandas as pd
import sys
import sqlite3
from sqlite3 import Error
from bisector import bisector
import re

def duplicationCheck(connection, table):
    conn = connection
    if conn:
        cursor = conn.cursor()
        data = []
        query = "SELECT DISTINCT date([Date]), [Cell Name], count([Cell Name])FROM {} GROUP BY [Cell Name] , date([Date]) having (count([Cell Name]) > 1) ORDER BY date([Date])".format(table)
        for row in cursor.execute(query):
            data.append(list(row))

        df = pd.DataFrame(data, columns=['Date', 'Cell Name', 'Duplicate Count'])
        
        if not df.empty:
            print("--- Duplicated entry found [Exported to duplicate.csv] ---")
            df.to_csv("duplicate.csv", index=False)
        else:
            print("--- Data clean, no duplicated found ---")

def duplicateDelete(connection, table):
    conn = connection
    if conn:
        cursor = conn.cursor()
        data = []
        query = "DELETE FROM {0} WHERE ID NOT IN ( SELECT MIN(ID) FROM {0} GROUP BY [Cell Name], date([Date]))".format(table)
        
        cursor.execute(query) # execute delete
        connection.commit()
        print("--- {} duplicated row deleted ----".format(cursor.rowcount))

def connectDatabase(databasename):
    """ create a database connection to a database that resides
        in memory
    """

    conn = None
    try:
        conn = sqlite3.connect(databasename)
        print("Connected database version: {}".format(sqlite3.version))

    except Error as e:
        print("Error when connecting to database. Make sure {} file in same directory".format(databasename))
        print(e)
    finally:
        if conn:
            return conn
        else:
            return False

def createTable(connection ,tablename, columns, datatypes):
    conn = connection
    if conn:
        cursor = conn.cursor()
        queryString = "CREATE TABLE {}(ID INTEGER PRIMARY KEY AUTOINCREMENT UNIQUE NOT NULL , {})".format(tablename,",".join(map(lambda x, y : x + " " + y, columns, datatypes)))
        
        print("Creating table {}".format(tablename))
        cursor.execute(queryString)

def convertType(type):
    if type == "<class 'datetime.datetime'>":
        return "DATE"
    elif type == "<class 'str'>":
        return "TEXT"
    elif type == "<class 'float'>":
        return "REAL"
    else:
        return "TEXT"

def inspectWorksheet(filename):
    dataTypeLine = None
    headerLine = None

    # Read excel first sheet, 100 column and first two row
    wb = load_workbook(filename, read_only=True)
    sheetname = wb.sheetnames[0]
    ws = wb[sheetname]
    for index , row in enumerate(ws.iter_rows(min_row=1, max_col=100, max_row=2)):
        line = [cell.value for cell in row if not cell.value == None]
        if not index == 0:
            dataTypeLine = [str(type(cell)) for cell in line]
        else:
            headerLine = [str(cell) for cell in line]
    
    convertion = map(lambda x, y: x + "," + convertType(y), headerLine, dataTypeLine)

    csv_format = open("{}_config.csv".format(sheetname),"w+")
    csv_format.write("\n".join(convertion))
    csv_format.close()

def dropTable(connection, tablename):
    cursor = connection.cursor()
    queryString = "DROP TABLE {}".format(tablename)
    cursor.execute(queryString)
    connection.commit()
    connection.execute("vacuum")

def valueConvert(c, d):
    if d == 'DATE' or d == "TEXT":
        return "{}".format(c)
    else:
        if c == '' or c == None:
            return 0
        else:
            try:
                float(c)
                return c
            except:
                return 0

def viewTableConfig():
    TABLE_CONFIG = "config.csv"
    configfile = open(TABLE_CONFIG,"r")
    columns = []
    datatypes = []
    for line in configfile.readlines():
        column , datatype = line.replace("\n","").replace(".","_").split(",")
        columns.append("[{}]".format(column))
        datatypes.append(datatype)
    
    return columns, datatypes

def updateTable(connection, database, excelfilepath):
    #LTE_KPI_FILE = "LTE_KPI_BH_NEW_Query_Result_20190927093645363.xlsx"

    columns , datatypes = viewTableConfig()
    wb = load_workbook(excelfilepath, read_only=True)

    conn = connection
    if conn:
        cursor = conn.cursor()
        sheetnames = wb.sheetnames
        insertRowCount = 0
        ws = wb[sheetnames[0]]

        print("Updating table {}".format(database))
        for index , row in enumerate(ws.iter_rows(min_row=1, max_col=len(columns), max_row=1000000)):
            line = [cell.value  for cell in row]
            if not (index == 0 or line[0] == None):
                insertString = "INSERT INTO {} ({}) VALUES ({})".format(database, ",".join(columns),",".join(["?"]*len(columns)))
                cursor.execute(insertString, tuple(map(valueConvert, line ,datatypes)))
                insertRowCount += cursor.rowcount
                sys.stdout.write("Update Success Count: {}\r".format(insertRowCount))
                sys.stdout.flush()
        sys.stdout.write("Update Success Count: {}\r\n".format(insertRowCount))
        sys.stdout.flush()
        conn.commit()


def main():

    if len(sys.argv) <= 1:
        print("""Available command:
delete          [Delete table from database]
compress        [Recalculate database size]
config          [View table config from config.csv file]
create          [Create table ]
duplicatecheck  [Check duplicate data]
duplicateremove [Remove duplicate data]
inspect         [Inspect excel sheet and generate table configuration]
update          [Update/append data to table]
bisector        [Generate BiSector Report]
help            [View help]
        """)
        return

    operation = sys.argv[1]
    conn = None
    if operation == 'help':
        print("""Available command:
delete          [Delete table from database]
compress        [Recalculate database size]
config          [View table config from config.csv file]
create          [Create table ]
duplicatecheck  [Check duplicate data]
duplicateremove [Remove duplicate data]
inspect         [Inspect excel sheet and generate table configuration]
update          [Update/append data to table]
bisector        [Generate BiSector Report]
help            [View help]""")

    if operation == 'delete': # delete database
        
        if len(sys.argv) <= 2:
            print("Invalid command. Missing table name arguments")
            print("""Usage: app.py [table name]""")
            return
        elif sys.argv[2] == "help":
            print("""Usage: app.py [table name]""")
            return
        conn = connectDatabase("maxis.db")
        table = sys.argv[2] # table to be deleted
        if conn:
            print("Deleting data {}".format(table))
            dropTable(conn, table)
    elif operation == 'duplicatecheck':
        conn = connectDatabase("maxis.db")
        table = sys.argv[2] # table to be checked
        if conn:
            duplicationCheck(conn, table)
    elif operation == 'duplicateremove':
        conn = connectDatabase("maxis.db")
        table = sys.argv[2]
        if conn:
            duplicateDelete(conn, table)

    elif operation == 'compress':
        conn = connectDatabase("maxis.db")
        if conn:
            print("Compressing databases")
            conn.execute("vacuum")
    elif operation == 'inspect':
        print("Inspecting sheet {}".format(sys.argv[2]))
        inspectWorksheet(sys.argv[2])
    elif operation == 'config':
        columns, datatypes = viewTableConfig()
        print("Type \t|Column\n{}".format("\n".join(map(lambda c, d: c + "\t|" + d, datatypes, columns))))
    elif operation == 'create':
        conn = connectDatabase("maxis.db")
        table = sys.argv[2] # table to be created
        columns, datatypes = viewTableConfig()
        createTable(conn, table, columns, datatypes)
    elif operation == "update":
        conn = connectDatabase("maxis.db")
        table = sys.argv[2] # table to be updated
        excelFile = sys.argv[3]
        updateTable(conn, table, excelFile)
    elif operation == "bisector":
        # default usage
        defaultUsage = """Usage: app.exe bisector [sitename] [bisector cellname 1] [bisector cellname 2] [outputfilename] [chart start date] [chart end date] [pre bisector date] [post bisector date]
------------------------------------------------------------------------------------------------
Example: app.exe ECODM ECODME_3 ECODME_C output.xlsx 2019-09-17 2019-09-18 2019-09-17 2019-09-25
------------------------------------------------------------------------------------------------
            """
        if len(sys.argv) <= 3 and sys.argv[2] == 'help':
            print(defaultUsage)
            return 
        
        if len(sys.argv) < 10:
            print("Invalid command. Missing arguments. ")
            print(defaultUsage)
            return 

        conn = connectDatabase("maxis.db")
        sitename = sys.argv[2] # "ECODM" 
        bisec1 =  sys.argv[3] # "ECODME_3"
        bisec2 =  sys.argv[4] # "ECODME_C"
        outputFilename =  sys.argv[5] # "output2.xlsx"

        if not re.match(r'\d{4}-\d{2}-\d{2}', sys.argv[6]):
            print("Invalid date: ", sys.argv[6])
            print(defaultUsage)
            return 
        chartStartDate =  sys.argv[6] # "2019-09-17"

        if not re.match(r'\d{4}-\d{2}-\d{2}', sys.argv[7]):
            print("Invalid date: ", sys.argv[7])
            print(defaultUsage)
            return 
        chartEndDate =  sys.argv[7] # "2019-09-18"

        if not re.match(r'\d{4}-\d{2}-\d{2}', sys.argv[8]):
            print("Invalid date: ", sys.argv[8])
            print(defaultUsage)
            return 
        preBisecDate =  sys.argv[8] # "2019-09-17"

        if not re.match(r'\d{4}-\d{2}-\d{2}', sys.argv[9]):
            print("Invalid date: ", sys.argv[9])
            print(defaultUsage)
            return 
        postBisecDate =  sys.argv[9] # "2019-09-25"

        bisectorReport = bisector(conn, sitename , bisec1, bisec2, outputFilename, chartStartDate, chartEndDate, preBisecDate, postBisecDate)
        bisectorReport.generateReport()

    if conn:
        conn.close()

if __name__ == '__main__':
    main()