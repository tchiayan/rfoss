import sqlite3
from openpyxl import load_workbook
import sys 
import re
import json

def connectDatabase(databasename):
    """ create a database connection to a database that resides
        in memory
    """

    conn = None
    try:
        conn = sqlite3.connect(databasename)
        print("Connected database version: {}".format(sqlite3.version))

    except Error as e:
        print("Error when connecting to database. Make sure {} file in same directory".format(databasename))
        print(e)
    finally:
        if conn:
            return conn
        else:
            return False

def findType(type):
    if type == "<class 'datetime.datetime'>":
        return "DATE"
    elif type == "<class 'str'>":
        return "TEXT"
    elif type == "<class 'float'>":
        return "REAL"
    else:
        return "TEXT"

def normalizeSqlName(name):
    # replace . or space with _
    _name = re.sub(r'[\s|.|%|+|?]', "_" , name.encode("ascii", 'ignore').decode())

    # replace () in the start or end with nothing
    _name = re.sub(r'[\(|\)]', "_", _name)
    _name = re.sub(r'^\_+|\_+$', "", _name)
    _name = re.sub(r'\_{2,}', "_", _name)
    _name = re.sub(r'-', "to", _name)
    _name = re.sub(r'\[\w+\]', "", _name)

    return re.sub(r'^\d+\_*',"", _name)

def updateTable(conn, tablename, excelfilepath):
    #LTE_KPI_FILE = "LTE_KPI_BH_NEW_Query_Result_20190927093645363.xlsx"

    #columns , datatypes = viewTableConfig()
    wb = load_workbook(excelfilepath, read_only=True)

    if conn:
        cursor = conn.cursor()
        sheetnames = wb.sheetnames
        insertRowCount = 0
        ws = wb[sheetnames[0]]

        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='{}'".format(tablename))
        
        if not cursor.fetchone():
            print('Table does not exists. Creating table.')

            headers = []
            # inspect ws first row for header
            for index, row in enumerate(ws.iter_rows(min_row=1, max_col=500, max_row=2)):
                line = [cell.value for cell in row if not cell.value == None]
                if index == 0:
                    headers = [normalizeSqlName(str(cell)) for cell in line]
                else:
                    columnType = [findType(str(type(cell))) for cell in line]

            createTableQuery = "CREATE TABLE {} (ID INTEGER PRIMARY KEY AUTOINCREMENT UNIQUE NOT NULL , {})".format(tablename,",".join(map(lambda x, y : x + " " + y, headers, columnType)))
            try:
                cursor.execute(createTableQuery)
            except Exception as e:
                print("Create table command error: ", createTableQuery)
                #print(createTableQuery)
                raise(e)
        else:
            # match database header and upload table header
            # alter database table if upload table header not yet exist on database header , provided existing database header not missing from upload table header
            cursor.execute("SELECT * FROM {} LIMIT 1".format(tablename))
            headers = list(map(lambda x: x[0], cursor.description))[1:]
            
            for index, row in enumerate(ws.iter_rows(min_row=1, max_col=500, max_row=2)):
                line = [cell.value for cell in row if not cell.value == None]
                if index == 0:
                    uploadHeader = [normalizeSqlName(str(cell)) for cell in line]
                else:
                    columnType = [findType(str(type(cell))) for cell in line]

            #print(uploadHeader)
            #print(headers)
            set_headers = set(headers)
            set_uploadHeader = set(uploadHeader)
            missing_header = list(set_uploadHeader-set_headers)
            if not len(missing_header) == 0:
                print(json.dumps({'error':'missing column'}))
                return

        # get total rowcount 
        if ws.max_row == 1:
            ws.reset_dimensions()
            ws.calculate_dimension(force=True)

        maxcount = ws.max_row
        '''hasnext = True
        batchcount = 0
        readbatch = 20000
        maxcount = 0
        while hasnext:
            currentCount = [0 if row[0].value == None else 1 for index , row in enumerate(ws.iter_rows(min_row=1+batchcount*readbatch, max_col=1, max_row=(1+batchcount)*readbatch))].count(1)
            maxcount += currentCount
            print("Current batch:" , currentCount, "Total batch", maxcount)
            if not currentCount == readbatch:
                hasnext = False
                print("Reading finish")
            else:
                #print("Reading next batch")
                batchcount += 1
            
        print(maxcount)
        print("Updating table {}".format(tablename))'''
        cursor.execute("SELECT * FROM {} LIMIT 1".format(tablename))
        headers = list(map(lambda x: x[0], cursor.description))[1:]

        #progress = 0
        #for index , row in enumerate(ws.iter_rows(min_row=1, max_col=len(headers), max_row=1000000)):
        for index , row in enumerate(ws.rows):
            line = [cell.value  for cell in row]
            if not (index == 0 or line[0] == None):
                insertString = "INSERT INTO {} ({}) VALUES ({})".format(tablename, ",".join(headers),",".join(["?"]*len(headers)))
                #print(insertString)
                #cursor.execute(insertString, tuple(map(valueConvert, line ,datatypes)))
                cursor.execute(insertString, tuple(line))
                insertRowCount += cursor.rowcount
                #if not progress == round(insertRowCount/float(maxcount),2):
                progress = round(insertRowCount/float(maxcount),2)
                print(json.dumps({'update_count':insertRowCount, 'progress':progress}))
                
                
        sys.stdout.write("Update Success Count: {}\r\n".format(insertRowCount))
        sys.stdout.flush()
        conn.commit()
    
if __name__ == "__main__":
    # API
    # upload dbpath tablename excelpath
    dbpath = sys.argv[2]
    tablename = sys.argv[3]
    excelpath = sys.argv[4]

    db_con = connectDatabase(dbpath)
    updateTable(db_con, tablename, excelpath)


